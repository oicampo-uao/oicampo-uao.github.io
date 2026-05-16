#!/usr/bin/env python3
"""
conversion_xlsx_to_md.py  —  v4.0  (UNIVERSAL)
Convierte cualquier RMF.xlsx con la estructura de plantilla BIV-300 a 9
archivos markdown para GitHub DHF, según ISO 14971 / ISO TR 24971.

Uso:
    python3 conversion_xlsx_to_md.py MiDispositivo_RMF.xlsx

Salida (carpeta ./rmf_output/):
    00_RMP.md  01_PHA.md  02_dFMEA.md  03_uFMEA.md  04_Risk_Evaluation.md
    05_Risk_Control.md  06_Overall_Residual_Risk.md  07_RMR.md
    08_Post_Production_Monitoring.md

Diseño universal:
    - Lee data_only=True  → obtiene valores calculados, no fórmulas
    - Detecta filas de encabezado buscando el texto esperado en col A,
      en lugar de asumir números de fila fijos.
    - Lee datos hasta encontrar la primera fila completamente vacía,
      por lo que el número de riesgos puede ser cualquiera.
    - Convierte #N/A y None a '—' (no inventa datos).
    - 00_RMP usa detección de sección por texto en col A.
"""

import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ Se requiere openpyxl.  Instala con:  pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES GENÉRICAS
# ─────────────────────────────────────────────────────────────────────────────

def fmt(v, sci=False):
    """Valor de celda → string limpio para Markdown."""
    if v is None:
        return '—'
    s = str(v).strip()
    if s in ('', '#N/A', 'None', '#VALUE!', '#REF!', '#NAME?'):
        return '—'
    if isinstance(v, float):
        return f"{v:.2E}" if sci else f"{v:g}"
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return s


def find_header_row(ws, first_col_value):
    """
    Busca la fila cuyo valor en col A coincide exactamente con first_col_value.
    Retorna el número de fila o None si no se encuentra.
    """
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() == str(first_col_value).strip():
            return r
    return None


def find_section_row(ws, section_text):
    """
    Busca una fila cuyo valor en col A empieza con section_text (mayúsculas).
    Útil para localizar bloques en 00_RMP (SEVERITY SCALE, etc.).
    """
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip().upper().startswith(section_text.upper()):
            return r
    return None


def read_table(ws, header_row):
    """
    Lee encabezados de `header_row` y datos desde la fila siguiente
    hasta la primera fila donde col A está vacía.
    Retorna (list[str] headers, list[dict] rows).
    """
    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        headers.append(str(h).strip() if h is not None else f"_col{c}")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        first = ws.cell(r, 1).value
        if first is None or str(first).strip() == '':
            break                       # primera fila vacía = fin de tabla
        row = {}
        for c, h in enumerate(headers, start=1):
            row[h] = ws.cell(r, c).value
        rows.append(row)
    return headers, rows


def md_table(col_labels, rows, getters):
    """
    Construye tabla Markdown.
      col_labels : lista de títulos visibles
      rows       : lista de dicts (del Excel)
      getters    : lista de (clave_dict, usar_notacion_sci)
    """
    sep = '|' + '|'.join(['---'] * len(col_labels)) + '|'
    lines = ['| ' + ' | '.join(col_labels) + ' |', sep]
    for row in rows:
        cells = [fmt(row.get(k), sci) for k, sci in getters]
        lines.append('| ' + ' | '.join(cells) + ' |')
    return '\n'.join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# GENERADOR 00_RMP
# ─────────────────────────────────────────────────────────────────────────────

def gen_00_rmp(ws):
    """
    Estructura de 00_RMP (detectada por texto en col A):
      Sección INFORMATION   → filas con pares (clave, valor) en cols A y B
      Sección SEVERITY SCALE → fila de encabezado + 5 filas S1-S5
      Sección PROBABILITY SCALE → fila de encabezado + 5 filas P1-P5
      Sección ACCEPTABILITY MATRIX → fila de subencabezado (S1-S5) + 5 filas de datos
    """

    # ── Metadatos ──────────────────────────────────────────────────────────
    meta = {}
    info_row = find_section_row(ws, 'INFORMATION')
    if info_row:
        for r in range(info_row + 1, ws.max_row + 1):
            k = ws.cell(r, 1).value
            v = ws.cell(r, 2).value
            if k is None or str(k).strip() == '':
                break
            meta[str(k).strip()] = v

    device       = fmt(meta.get('device_name'))
    manufacturer = fmt(meta.get('manufacturer'))
    version      = fmt(meta.get('version'))
    created      = fmt(meta.get('created_date'))
    responsible  = fmt(meta.get('responsible'))

    # ── Escala Severidad ───────────────────────────────────────────────────
    sev_hdr = find_section_row(ws, 'SEVERITY SCALE')
    sev_rows_md = []
    if sev_hdr:
        for r in range(sev_hdr + 2, ws.max_row + 1):   # +2 salta cabecera
            a = ws.cell(r, 1).value
            if a is None or str(a).strip() == '': break
            vals = [fmt(ws.cell(r,c).value) for c in range(1,4)]   # solo 3 cols
            sev_rows_md.append('| ' + ' | '.join(vals) + ' |')

    # ── Escala Probabilidad ────────────────────────────────────────────────
    prob_hdr = find_section_row(ws, 'PROBABILITY SCALE')
    prob_rows_md = []
    if prob_hdr:
        for r in range(prob_hdr + 2, ws.max_row + 1):  # +2 salta cabecera
            a = ws.cell(r, 1).value
            if a is None or str(a).strip() == '': break
            vals = [fmt(ws.cell(r,c).value) for c in range(1,5)]   # solo 4 cols
            prob_rows_md.append('| ' + ' | '.join(vals) + ' |')

    # ── Matriz de Aceptabilidad ────────────────────────────────────────────
    mat_hdr = find_section_row(ws, 'ACCEPTABILITY MATRIX')
    matrix_rows_md = []
    if mat_hdr:
        # La fila de datos empieza donde col A = 'P' (primer nivel de probabilidad)
        p_start = find_header_row(ws, 'P')
        if p_start is None:
            # Fallback: buscar la primera fila después del encabezado de matriz
            # cuyo col A no está vacía
            for r in range(mat_hdr + 1, ws.max_row + 1):
                if ws.cell(r, 1).value is not None:
                    p_start = r
                    break
        if p_start:
            for r in range(p_start, ws.max_row + 1):
                a = ws.cell(r, 1).value
                if a is None or str(a).strip() == '':
                    break
                p_nivel = fmt(ws.cell(r, 2).value)
                vals = [fmt(ws.cell(r, c).value) for c in range(3, 8)]
                matrix_rows_md.append(f"| **P{p_nivel}** | " + " | ".join(vals) + " |")

    sev_str  = '\n'.join(sev_rows_md)  if sev_rows_md  else '*(sin datos)*'
    prob_str = '\n'.join(prob_rows_md) if prob_rows_md else '*(sin datos)*'
    mat_str  = '\n'.join(matrix_rows_md) if matrix_rows_md else '*(sin datos)*'

    return f"""# Risk Management Plan (RMP)
**Documento:** RMP-001  
**Versión:** {version}  
**Fecha:** {created}  
**Responsable:** {responsible}

---

## 1. Información del Dispositivo

| Campo | Valor |
|---|---|
| Dispositivo | {device} |
| Fabricante / Institución | {manufacturer} |
| Versión | {version} |
| Fecha | {created} |
| Responsable | {responsible} |

---

## 2. Escala de Severidad

| Nivel | Descriptor | Descripción |
|---|---|---|
{sev_str}

---

## 3. Escala de Probabilidad

| Nivel | Rango | Descriptor | Interpretación |
|---|---|---|---|
{prob_str}

---

## 4. Matriz de Aceptabilidad (S × P)

|  | **S1** | **S2** | **S3** | **S4** | **S5** |
|---|:---:|:---:|:---:|:---:|:---:|
{mat_str}

> **✅ Aceptable** · **ALARP** = reducir tanto como sea razonablemente practicable · **❌ Inaceptable**

---
_Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_
"""


# ─────────────────────────────────────────────────────────────────────────────
# GENERADORES DE TABLAS VARIABLES (01–05)
# ─────────────────────────────────────────────────────────────────────────────

def gen_01_pha(ws):
    """Encabezados detectados buscando 'ID_Hazard' en col A."""
    hdr = find_header_row(ws, 'ID_Hazard')
    if hdr is None:
        return "# PHA\n\n*(Encabezado ID_Hazard no encontrado)*\n"
    _, rows = read_table(ws, hdr)

    table = md_table(
        ['ID', 'Peligro', 'Descripción', 'Severidad', 'Fuente'],
        rows,
        [('ID_Hazard', False), ('Hazard_Name', False), ('Description', False),
         ('Severity', False), ('Source', False)]
    )
    return f"""# Preliminary Hazard Analysis (PHA)
**Documento:** PHA-001

---

## Peligros Identificados

{table}

---

**Total peligros identificados: {len(rows)}**

_Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_
"""


def gen_02_dfmea(ws):
    """
    Encabezados detectados buscando 'ID_Hazard' en col A.
    Las 19 columnas reales son:
    ID_Hazard Domain Failure_Mode Effect_Patient S P1_method P1 Mechanism_P2
    P2 P3 P_daño Nivel_P RI_inicial Control_Desc P2_residual P3_residual
    P_daño_residual Nivel_P_residual RR_residual
    """
    hdr = find_header_row(ws, 'ID_Hazard')
    if hdr is None:
        return "# dFMEA\n\n*(Encabezado ID_Hazard no encontrado)*\n"
    _, rows = read_table(ws, hdr)

    t1 = md_table(
        ['ID', 'Dominio', 'Modo de Fallo', 'Efecto Paciente', 'S',
         'P1_método', 'P1', 'Mecan. P2', 'P2', 'P3', 'P_daño', 'Nivel_P', 'RI_inicial'],
        rows,
        [('ID_Hazard',False), ('Domain',False), ('Failure_Mode',False),
         ('Effect_Patient',False), ('S',False),
         ('P1_method',False), ('P1',True), ('Mechanism_P2',False),
         ('P2',False), ('P3',False),
         ('P_daño',True), ('Nivel_P',False), ('RI_inicial',False)]
    )
    t2 = md_table(
        ['ID', 'Control', 'P2_res', 'P3_res', 'P_daño_res', 'Nivel_P_res', 'RR_residual'],
        rows,
        [('ID_Hazard',False), ('Control_Desc',False),
         ('P2_residual',False), ('P3_residual',False),
         ('P_daño_residual',True), ('Nivel_P_residual',False), ('RR_residual',False)]
    )
    return f"""# Design FMEA (dFMEA)
**Documento:** DFMEA-001  
**Subsistemas:** Mecánico · Electrónico · Software

> P_daño, Nivel_P, RI_inicial, P_daño_residual, Nivel_P_residual y RR_residual son
> calculados automáticamente por la plantilla (INDEX/MATCH sobre MatrizAceptabilidad).

---

## Fallos y Riesgo Inicial

{t1}

---

## Control y Riesgo Residual

{t2}

---
_Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_
"""


def gen_03_ufmea(ws):
    """
    Columnas reales:
    ID_Hazard User_Error Context S P1_method P1 Protections P2 P3
    P_daño Nivel_P RI_inicial Control_Desc P2_residual P3_residual
    P_daño_residual Nivel_P_residual RR_residual
    """
    hdr = find_header_row(ws, 'ID_Hazard')
    if hdr is None:
        return "# uFMEA\n\n*(Encabezado ID_Hazard no encontrado)*\n"
    _, rows = read_table(ws, hdr)

    t1 = md_table(
        ['ID', 'Error de Uso', 'Contexto', 'S',
         'P1_método', 'P1', 'Protecciones', 'P2', 'P3', 'P_daño', 'Nivel_P', 'RI_inicial'],
        rows,
        [('ID_Hazard',False), ('User_Error',False), ('Context',False), ('S',False),
         ('P1_method',False), ('P1',False), ('Protections',False),
         ('P2',False), ('P3',False),
         ('P_daño',True), ('Nivel_P',False), ('RI_inicial',False)]
    )
    t2 = md_table(
        ['ID', 'Control', 'P2_res', 'P3_res', 'P_daño_res', 'Nivel_P_res', 'RR_residual'],
        rows,
        [('ID_Hazard',False), ('Control_Desc',False),
         ('P2_residual',False), ('P3_residual',False),
         ('P_daño_residual',True), ('Nivel_P_residual',False), ('RR_residual',False)]
    )
    return f"""# Usability FMEA (uFMEA)
**Documento:** UFMEA-001  
**Referencia:** IEC 62366-1:2015

> P_daño, Nivel_P, RI_inicial, P_daño_residual, Nivel_P_residual y RR_residual son
> calculados automáticamente por la plantilla.

---

## Errores de Uso y Riesgo Inicial

{t1}

---

## Control y Riesgo Residual

{t2}

---
_Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_
"""


def gen_04_risk_eval(ws):
    """
    Columnas reales:
    ID_Hazard Domain Hazard_Desc S Nivel_P_i RI_i Nivel_P_r RR_r
    """
    hdr = find_header_row(ws, 'ID_Hazard')
    if hdr is None:
        return "# Risk Evaluation\n\n*(Encabezado ID_Hazard no encontrado)*\n"
    _, rows = read_table(ws, hdr)

    table = md_table(
        ['ID', 'Dominio', 'Peligro / Descripción', 'S',
         'Nivel_P inicial', 'RI inicial', 'Nivel_P residual', 'RR residual'],
        rows,
        [('ID_Hazard',False), ('Domain',False), ('Hazard_Desc',False), ('S',False),
         ('Nivel_P_i',False), ('RI_i',False), ('Nivel_P_r',False), ('RR_r',False)]
    )
    return f"""# Risk Evaluation — Consolidada
**Documento:** RE-001

> Valores calculados automáticamente via VLOOKUP desde 02_dFMEA y 03_uFMEA,
> usando la MatrizAceptabilidad definida en 00_RMP.

---

{table}

---
_Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_
"""


def gen_05_risk_control(ws):
    """
    Columnas reales:
    ID_Control ID_Hazard S Nivel_P_i RI_i Control_Type Control_Desc
    Verification Nivel_P_r RR_r Status
    """
    hdr = find_header_row(ws, 'ID_Control')
    if hdr is None:
        return "# Risk Control\n\n*(Encabezado ID_Control no encontrado)*\n"
    _, rows = read_table(ws, hdr)

    table = md_table(
        ['ID Control', 'ID Hazard', 'S', 'Nivel_P_i', 'RI_i',
         'Tipo', 'Descripción del Control', 'Verificación',
         'Nivel_P_r', 'RR_r', 'Estado'],
        rows,
        [('ID_Control',False), ('ID_Hazard',False), ('S',False),
         ('Nivel_P_i',False), ('RI_i',False),
         ('Control_Type',False), ('Control_Desc',False), ('Verification',False),
         ('Nivel_P_r',False), ('RR_r',False), ('Status',False)]
    )
    return f"""# Risk Control
**Documento:** RC-001  
**Referencia:** ISO 14971:2019 Sección 7

---

## Jerarquía de Control (ISO 14971 §7.1)

1. **Inherently safe design** — diseño inherentemente seguro
2. **Protective measure** — barreras, interlocks, alarmas
3. **Safety information** — IFU, etiquetas, entrenamiento (última opción)

---

## Controles Implementados

{table}

---
_Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_
"""


# ─────────────────────────────────────────────────────────────────────────────
# GENERADORES DE HOJAS FIJAS (06–08)
# ─────────────────────────────────────────────────────────────────────────────

def gen_06_overall_rr(ws):
    """Columnas: Métrica | Inicial | Residual. Encabezado buscado por 'Métrica'."""
    hdr = find_header_row(ws, 'Métrica')
    if hdr is None:
        return "# Overall Residual Risk\n\n*(Encabezado Métrica no encontrado)*\n"
    _, rows = read_table(ws, hdr)

    table = md_table(
        ['Métrica', 'Inicial', 'Residual'],
        rows,
        [('Métrica',False), ('Inicial',False), ('Residual',False)]
    )
    return f"""# Overall Residual Risk Evaluation
**Documento:** ORR-001  
**Referencia:** ISO 14971:2019 Sección 8

---

## Resumen de Riesgos

{table}

---

## Conclusión del Panel

- ¿Algún riesgo residual inaceptable? → Ver tabla (debe ser 0)
- ¿Los beneficios superan los riesgos residuales? → Documentar aquí
- ¿Perfil comparable al estado del arte? → Documentar aquí

---
_Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_
"""


def gen_07_rmr(ws):
    """Columnas: Objetivo | Criterio | Resultado | Evidencia."""
    hdr = find_header_row(ws, 'Objetivo')
    if hdr is None:
        return "# RMR\n\n*(Encabezado Objetivo no encontrado)*\n"
    _, rows = read_table(ws, hdr)

    table = md_table(
        ['Objetivo', 'Criterio', 'Resultado', 'Evidencia'],
        rows,
        [('Objetivo',False), ('Criterio',False), ('Resultado',False), ('Evidencia',False)]
    )
    return f"""# Risk Management Review (RMR)
**Documento:** RMR-001  
**Referencia:** ISO 14971:2019 Sección 9

> Esta revisión se realiza DESPUÉS de verificar todas las medidas de control
> y ANTES del lanzamiento comercial.

---

## Verificación del Plan de Gestión de Riesgos

{table}

---

## Firma y Aprobación

| Rol | Nombre | Firma | Fecha |
|---|---|---|---|
| Risk Management Owner | | | |
| Director de Calidad | | | |
| Experto Clínico Independiente | | | |
| Director de Ingeniería | | | |

---
_Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_
"""


def gen_08_postprod(ws):
    """Columnas: Actividad | Indicador | Frecuencia | Responsable."""
    hdr = find_header_row(ws, 'Actividad')
    if hdr is None:
        return "# Post-Production Monitoring\n\n*(Encabezado Actividad no encontrado)*\n"
    _, rows = read_table(ws, hdr)

    table = md_table(
        ['Actividad', 'Indicador', 'Frecuencia', 'Responsable'],
        rows,
        [('Actividad',False), ('Indicador',False),
         ('Frecuencia',False), ('Responsable',False)]
    )
    return f"""# Post-Production Monitoring
**Documento:** PP-001  
**Referencia:** ISO 14971:2019 Sección 10

---

## Plan de Vigilancia

{table}

---

## Criterios de Acción (Triggers)

- Evento adverso grave (S4-S5) → acción inmediata (24-72h)
- ≥ 3 eventos del mismo modo de fallo en un semestre → revisión del control
- Cambio en estándares aplicables → actualización del RMF

---
_Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}_
"""


# ─────────────────────────────────────────────────────────────────────────────
# CLASE PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

class XlsxToMarkdown:

    # (nombre_archivo_salida, nombre_hoja_excel, función_generadora)
    GENERATORS = [
        ('00_RMP.md',                        '00_RMP',          gen_00_rmp),
        ('01_PHA.md',                        '01_PHA',          gen_01_pha),
        ('02_dFMEA.md',                      '02_dFMEA',        gen_02_dfmea),
        ('03_uFMEA.md',                      '03_uFMEA',        gen_03_ufmea),
        ('04_Risk_Evaluation.md',            '04_Risk_Eval',    gen_04_risk_eval),
        ('05_Risk_Control.md',               '05_Risk_Control', gen_05_risk_control),
        ('06_Overall_Residual_Risk.md',      '06_Overall_RR',   gen_06_overall_rr),
        ('07_RMR.md',                        '07_RMR',          gen_07_rmr),
        ('08_Post_Production_Monitoring.md', '08_PostProd',     gen_08_postprod),
    ]

    def __init__(self, xlsx_file):
        self.xlsx_file = xlsx_file
        try:
            self.wb = load_workbook(xlsx_file, data_only=True)
        except FileNotFoundError:
            print(f"❌ Archivo no encontrado: {xlsx_file}")
            sys.exit(1)
        self.output_dir = Path("./rmf_output")
        self.output_dir.mkdir(exist_ok=True)

    def run(self):
        print("\n" + "="*60)
        print("  CONVERSIÓN: RMF.xlsx → Markdown  v4.0 (Universal)")
        print("="*60)
        print(f"\n📥 Archivo: {self.xlsx_file}")
        print(f"📋 Hojas:   {', '.join(self.wb.sheetnames)}\n")

        ok = err = 0
        for filename, sheet_name, generator in self.GENERATORS:
            if sheet_name not in self.wb.sheetnames:
                print(f"  ⚠️  {filename:<48} — hoja '{sheet_name}' no encontrada")
                err += 1
                continue
            try:
                content = generator(self.wb[sheet_name])
                (self.output_dir / filename).write_text(content, encoding='utf-8')
                print(f"  ✅ {filename}")
                ok += 1
            except Exception as e:
                print(f"  ❌ {filename:<48} — {e}")
                import traceback; traceback.print_exc()
                err += 1

        print(f"\n📁 Salida: {self.output_dir.resolve()}/")
        print(f"   {ok} generados · {err} errores")
        print("\n" + "="*60)
        print("  ✅ COMPLETADO" if err == 0 else f"  ⚠️  COMPLETADO CON {err} ERRORES")
        print("="*60 + "\n")
        print("\n📝 Próximos pasos:")
        print("   1. Revisa los archivos generados en ./rmf_output/")
        print("   2. Si falta algo, ajusta en Excel")
        print("   3. Vuelve a ejecutar este script")
        print("   4. Una vez listos, copia los .md a tu DHF en GitHub")
        print("\n")


def main():
    if len(sys.argv) < 2:
        print("Uso:     python3 conversion_xlsx_to_md.py <archivo.xlsx>")
        print("Ejemplo: python3 conversion_xlsx_to_md.py BIV-300_RMF.xlsx")
        sys.exit(1)
    XlsxToMarkdown(sys.argv[1]).run()


if __name__ == "__main__":
    main()
